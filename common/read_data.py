from configparser import ConfigParser


class MyConfigParser(ConfigParser):
    # 重写 configparser 中的 optionxform 函数，解决 .ini 文件中的 键option 自动转为小写的问题
    def __init__(self, defaults=None):
        ConfigParser.__init__(self, defaults=defaults)

    def optionxform(self, option_str):
        return option_str


class ReadFileData:

    def __init__(self):
        pass

    @staticmethod
    def load_yaml(file_path):
        import yaml
        with open(file_path, encoding='utf-8') as f:
            data = yaml.safe_load(f)
        return data

    @staticmethod
    def load_json(file_path):
        import json
        with open(file_path, encoding='utf-8') as f:
            data = json.load(f)
        return data

    @staticmethod
    def load_ini(file_path):
        config = MyConfigParser()
        config.read(file_path, encoding="UTF-8")
        data = dict(config._sections)
        return data

    @staticmethod
    def load_excel(file_path, sheet_name=None):
        from openpyxl import load_workbook
        work_book = load_workbook(file_path)
        if sheet_name in work_book.sheetnames:
            sheet = work_book[sheet_name]
        else:
            sheet = work_book.active
        all_data = []
        titles = [cell.value for cell in list(sheet.rows)[0]]
        for item in list(sheet.rows)[1:]:
            value_list = [cell.value for cell in item]
            dic = dict(zip(titles, value_list))
            all_data.append(dic)
        return all_data


data = ReadFileData()

if __name__ == '__main__':
    print()
